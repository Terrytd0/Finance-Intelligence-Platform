import datetime
from pathlib import Path

from openpyxl import Workbook

from src.ingestion.pipeline import run_pipeline
from src.ingestion.reader import read_rows

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
SAMPLE_DIR = DATA_DIR / "sample"
RAW_DIR = DATA_DIR / "raw"


def test_valid_transactions_pass_with_no_errors():
    result = run_pipeline(SAMPLE_DIR / "valid_transactions.csv")

    assert len(result.transactions) == 4
    assert result.errors == []


def test_missing_values_are_rejected():
    result = run_pipeline(SAMPLE_DIR / "missing_values.csv")

    assert len(result.transactions) == 1
    rules = {error.rule for error in result.errors}
    assert rules == {"required_field"}


def test_invalid_data_is_rejected():
    result = run_pipeline(SAMPLE_DIR / "invalid_data.csv")

    assert result.transactions == []
    rules = {error.rule for error in result.errors}
    assert rules == {"invalid_date", "invalid_amount", "invalid_category"}


def test_duplicate_transaction_id_is_rejected_but_first_occurrence_kept():
    result = run_pipeline(SAMPLE_DIR / "duplicate_transactions.csv")

    assert [t.transaction_id for t in result.transactions] == ["T2001", "T2002"]
    assert len(result.errors) == 1
    assert result.errors[0].rule == "duplicate_transaction_id"


def test_xlsx_and_csv_sources_produce_equivalent_transactions():
    # transactions.xlsx stores dates/amounts as native Excel types, not
    # text, so a raw dict-for-dict comparison of read_rows() output isn't
    # meaningful (e.g. a numeric 125000.00 cell reads back as 125000, same
    # as real Excel). What has to match is the pipeline's parsed output.
    csv_result = run_pipeline(RAW_DIR / "transactions.csv")
    xlsx_result = run_pipeline(RAW_DIR / "transactions.xlsx")

    assert xlsx_result.transactions == csv_result.transactions
    assert xlsx_result.errors == csv_result.errors == []


def test_xlsx_reader_skips_hidden_sheets(tmp_path):
    workbook = Workbook()
    hidden = workbook.active
    hidden.title = "helper"
    hidden.sheet_state = "hidden"
    hidden.append(["should", "not", "appear"])

    visible = workbook.create_sheet("transactions")
    visible.append(["transaction_id", "client_id", "date", "category", "description", "amount", "currency"])
    visible.append(["T9001", "C001", datetime.date(2026, 1, 5), "revenue", "", 100, "USD"])

    path = tmp_path / "hidden_sheet.xlsx"
    workbook.save(path)

    rows = read_rows(path)

    assert len(rows) == 1
    assert rows[0]["transaction_id"] == "T9001"

def test_currency_is_normalized_to_uppercase():
    result = run_pipeline(SAMPLE_DIR / "lowercase_currency.csv")

    assert result.errors == []
    assert result.transactions[0].currency == "USD"

def test_cleaner_trims_whitespace():
    result = run_pipeline(SAMPLE_DIR / "whitespace_values.csv")

    transaction = result.transactions[0]

    assert transaction.description == "Salary payment"
    assert transaction.currency == "USD"

def test_xlsx_reader_fills_merged_cell_values(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "transactions"
    sheet.append(["transaction_id", "client_id", "date", "category", "description", "amount", "currency"])
    sheet.append(["T9001", "C001", datetime.date(2026, 1, 5), "revenue", "shared note", 100, "USD"])
    sheet.append(["T9002", "C001", datetime.date(2026, 1, 5), "expense", "shared note", -50, "USD"])
    # Merge the description column across both data rows; openpyxl only
    # stores the value on the top-left cell of a merged range.
    sheet.merge_cells("E2:E3")

    path = tmp_path / "merged_cells.xlsx"
    workbook.save(path)

    rows = read_rows(path)

    assert rows[0]["description"] == "shared note"
    assert rows[1]["description"] == "shared note"
