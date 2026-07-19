"""
Read supported source files into a normalized in-memory representation.

Returns:
    list[dict[str, str]]

Every downstream stage operates on this common representation, regardless
of whether the original source was CSV or Excel.
"""

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

def _read_csv(path: Path) -> list[dict[str, str]]:
    with open(path, newline="", encoding="utf-8") as f:
        return [dict(row) for row in csv.DictReader(f)]


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    # data_only=True: prefer a formula's last-cached computed value over its
    # formula text, since openpyxl has no calculation engine of its own.
    workbook = load_workbook(path, data_only=True)
    sheet = _first_visible_sheet(workbook)
    if sheet is None:
        return []

    rows = list(sheet.iter_rows())

    try: 
        if not rows:
            return []

        merged_values = _merged_cell_values(sheet)
        header = [_cell_str(cell.value) for cell in rows[0]]

        data_rows = []
        for row in rows[1:]:
            if all(cell.value is None for cell in row):
                continue
            values = [
                _cell_str(merged_values.get(cell.coordinate, cell.value))
                for cell in row
            ]
            data_rows.append(
                {header[i]: (values[i] if i < len(values) else "") for i in range(len(header))}
            )
        return data_rows
    finally:
        workbook.close()

def _first_visible_sheet(workbook):
    """Pick the first non-hidden worksheet. Real exports often carry
    hidden helper/calculation sheets alongside the actual data sheet."""
    for sheet in workbook.worksheets:
        if sheet.sheet_state == "visible":
            return sheet
    return None


def _merged_cell_values(sheet) -> dict[str, object]:
    """Map every coordinate inside a merged range to that range's value.
    openpyxl only stores the value on the merged range's top-left cell;
    every other cell in the range reads back as None."""
    values: dict[str, object] = {}
    for merged_range in sheet.merged_cells.ranges:
        anchor_value = sheet.cell(
            row=merged_range.min_row, column=merged_range.min_col
        ).value
        for row in sheet.iter_rows(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                values[cell.coordinate] = anchor_value
    return values


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)

READERS = {
    ".csv": _read_csv,
    ".xlsx": _read_xlsx,
}

def read_rows(path: str | Path) -> list[dict[str, str]]:
    path = Path(path)

    reader = READERS.get(path.suffix.lower())

    if reader is None:
        raise ValueError(f"Unsupported file type: {path.suffix}")

    return reader(path)